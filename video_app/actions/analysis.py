import shutil
from collections import defaultdict
from datetime import datetime

from video_app.actions.common import 上报进度, 缩放进度, 返回成功
from video_app.core.utils import (
    安全目标文件,
    参数校验错误,
    拆分关键词,
    捕获日志,
    查找视频文件,
    校验文件,
    校验目录,
    校验输出路径,
)


def _parse_video_records(source, progress=None):
    records = []
    files = 查找视频文件(source, extensions={".mp4"})
    for index, file_path in enumerate(files, start=1):
        stem = file_path.stem
        if "++=" not in stem:
            print(f"跳过，文件名不含分隔符 ++=: {file_path.name}")
            continue
        parts = [item.strip() for item in stem.split("++=") if item.strip()]
        if len(parts) < 3:
            print(f"跳过，字段不足: {file_path.name}")
            continue
        up_name = parts[0]
        publish_text = parts[1]
        duration_text = parts[-1]
        try:
            duration_seconds = int(duration_text)
        except ValueError:
            duration_seconds = 0
        date_candidates = [
            publish_text.replace("年", "-").replace("月", "-").replace("日", ""),
            publish_text.replace(".", "-").replace("/", "-"),
        ]
        publish_date = None
        for item in date_candidates:
            try:
                publish_date = datetime.strptime(item, "%Y-%m-%d")
                break
            except ValueError:
                continue
        if publish_date is None:
            print(f"跳过，日期无法解析: {file_path.name}")
            continue
        records.append(
            {
                "up_name": up_name,
                "publish_date": publish_date,
                "publish_date_text": publish_text,
                "year": publish_date.year,
                "duration_seconds": duration_seconds,
                "duration_category": _analysis_duration_category(duration_seconds),
                "file_name": file_path.name,
                "full_path": str(file_path),
            }
        )
        上报进度(progress, 缩放进度(index, len(files), 8, 34), "解析文件名", f"已解析 {index}/{len(files)}")
    return records


def _analysis_duration_category(seconds):
    if seconds <= 0:
        return "未知时长"
    if seconds <= 30:
        return "短视频(0-30秒)"
    if seconds <= 60:
        return "中短视频(30-60秒)"
    if seconds <= 180:
        return "中视频(60-180秒)"
    if seconds <= 300:
        return "长视频(180-300秒)"
    return "超长视频(300+秒)"


def UP主统计分析(params, progress=None):
    source = 校验目录(params.get("source_dir"), "分析目录")
    output_excel = 校验输出路径(params.get("output_excel"), "输出 Excel", required=False)

    def runner():
        import pandas as pd

        上报进度(progress, 5, "准备中", "正在解析视频文件名。")
        records = _parse_video_records(source, progress=progress)
        print(f"有效记录数: {len(records)}")
        if not records:
            raise 参数校验错误("未找到符合命名规则的 MP4 文件")
        上报进度(progress, 40, "统计中", "正在汇总 UP 主维度统计。")

        up_summary = defaultdict(
            lambda: {"total_videos": 0, "latest_date": None, "latest_date_text": "", "years": set(), "total_duration": 0}
        )
        yearly_stats = defaultdict(lambda: defaultdict(int))
        duration_stats = defaultdict(lambda: defaultdict(int))

        for item in records:
            up = up_summary[item["up_name"]]
            up["total_videos"] += 1
            up["total_duration"] += item["duration_seconds"]
            up["years"].add(item["year"])
            if up["latest_date"] is None or item["publish_date"] > up["latest_date"]:
                up["latest_date"] = item["publish_date"]
                up["latest_date_text"] = item["publish_date_text"]
            yearly_stats[item["up_name"]][item["year"]] += 1
            duration_stats[item["up_name"]][item["duration_category"]] += 1

        上报进度(progress, 60, "生成表格", "正在生成 Excel 数据表。")
        summary_rows = []
        for up_name, stats in sorted(up_summary.items(), key=lambda x: x[1]["total_videos"], reverse=True):
            avg_duration = stats["total_duration"] / stats["total_videos"] if stats["total_videos"] else 0
            summary_rows.append(
                {
                    "UP主名称": up_name,
                    "视频总数": stats["total_videos"],
                    "最新发布时间": stats["latest_date_text"],
                    "总时长秒数": stats["total_duration"],
                    "平均时长秒数": round(avg_duration, 2),
                    "活跃年份数": len(stats["years"]),
                    "活跃年份": ", ".join(str(year) for year in sorted(stats["years"])),
                }
            )
        summary_df = pd.DataFrame(summary_rows)

        all_years = sorted({year for year_map in yearly_stats.values() for year in year_map})
        yearly_rows = []
        for up_name in sorted(yearly_stats):
            row = {"UP主名称": up_name}
            total = 0
            for year in all_years:
                count = yearly_stats[up_name].get(year, 0)
                row[str(year)] = count
                total += count
            row["总计"] = total
            yearly_rows.append(row)
        yearly_df = pd.DataFrame(yearly_rows)

        duration_order = ["短视频(0-30秒)", "中短视频(30-60秒)", "中视频(60-180秒)", "长视频(180-300秒)", "超长视频(300+秒)", "未知时长"]
        duration_rows = []
        for up_name in sorted(duration_stats):
            row = {"UP主名称": up_name}
            total = 0
            for category in duration_order:
                count = duration_stats[up_name].get(category, 0)
                row[category] = count
                total += count
            row["总计"] = total
            duration_rows.append(row)
        duration_df = pd.DataFrame(duration_rows)

        global_rows = []
        total_records = len(records)
        for category in duration_order:
            count = sum(1 for item in records if item["duration_category"] == category)
            global_rows.append({"时长分类": category, "视频数量": count, "占比": f"{(count / total_records * 100):.1f}%"})
        global_df = pd.DataFrame(global_rows)

        output_path = output_excel if output_excel else source / f"UP主统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        上报进度(progress, 86, "写入文件", "正在写入 Excel 文件。")
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            summary_df.to_excel(writer, sheet_name="UP主概览", index=False)
            yearly_df.to_excel(writer, sheet_name="年度统计", index=False)
            duration_df.to_excel(writer, sheet_name="时长统计", index=False)
            global_df.to_excel(writer, sheet_name="全局时长分布", index=False)
        print(f"统计报表已输出: {output_path}")
        上报进度(progress, 100, "已完成", "统计分析完成。")
        return {"record_count": len(records), "output_path": str(output_path)}

    result, logs = 捕获日志(runner)
    return 返回成功("UP 主统计分析已完成", logs, result)


def 关键词整理(params, progress=None):
    source = 校验目录(params.get("source_dir"), "源目录")
    target = 校验目录(params.get("target_dir"), "目标目录")
    keywords = 拆分关键词(params.get("keywords"))
    if not keywords:
        raise 参数校验错误("请至少输入一个关键词")

    def runner():
        files = 查找视频文件(source)
        上报进度(progress, 8, "准备中", f"共找到 {len(files)} 个视频文件。")
        stats = {keyword: 0 for keyword in keywords}
        moved = 0
        for index, file_path in enumerate(files, start=1):
            lower_name = file_path.name.lower()
            for keyword in keywords:
                if keyword.lower() in lower_name:
                    safe_target = 安全目标文件(target / file_path.name)
                    shutil.move(str(file_path), str(safe_target))
                    stats[keyword] += 1
                    moved += 1
                    print(f"[{keyword}] 已移动: {file_path.name}")
                    break
            上报进度(progress, 缩放进度(index, len(files), 8, 100), "整理关键词", f"已处理 {index}/{len(files)}")
        return {"keyword_stats": stats, "moved_count": moved}

    result, logs = 捕获日志(runner)
    return 返回成功("关键词整理已完成", logs, result)


def Excel映射重命名(params, progress=None):
    folder_path = 校验目录(params.get("folder_path"), "视频目录")
    mapping_file = 校验文件(params.get("mapping_file"), "映射 Excel", required=False)
    output_excel = 校验输出路径(params.get("output_excel"), "输出报告 Excel", required=True)
    actually_rename = bool(params.get("actually_rename", False))
    backup_original = bool(params.get("backup_original", False))

    def runner():
        from openpyxl import Workbook, load_workbook

        上报进度(progress, 5, "准备中", "正在读取映射文件。")
        mapping = {}
        if mapping_file:
            wb = load_workbook(mapping_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
                if not row or row[1] is None or row[2] is None:
                    continue
                original_name = str(row[1]).strip()
                new_name = str(row[2]).strip()
                if original_name.lower().endswith(".mp4"):
                    original_name = original_name[:-4]
                if new_name.lower().endswith(".mp4"):
                    new_name = new_name[:-4]
                mapping[original_name] = new_name
            wb.close()
            print(f"读取映射关系数量: {len(mapping)}")
        backup_dir = None
        if actually_rename and backup_original:
            backup_dir = folder_path / "原文件备份"
            backup_dir.mkdir(exist_ok=True)

        processed = []
        files = list(folder_path.rglob("*.mp4"))
        上报进度(progress, 18, "扫描完成", f"共找到 {len(files)} 个 MP4 文件。")
        for index, file_path in enumerate(files, start=1):
            current_stem = file_path.stem
            record = {
                "原始文件名": file_path.name,
                "显示名称": current_stem,
                "状态": "无映射",
                "结果路径": str(file_path),
                "备注": "",
            }
            if current_stem in mapping:
                target_name = f"{mapping[current_stem]}.mp4"
                target_path = 安全目标文件(file_path.with_name(target_name))
                record["显示名称"] = mapping[current_stem]
                record["结果路径"] = str(target_path)
                if actually_rename:
                    if backup_dir:
                        shutil.copy2(str(file_path), str(backup_dir / file_path.name))
                    file_path.rename(target_path)
                    record["状态"] = "已重命名"
                else:
                    record["状态"] = "待重命名"
            processed.append(record)
            上报进度(progress, 缩放进度(index, len(files), 18, 82), "处理映射", f"已处理 {index}/{len(files)}")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "处理报告"
        headers = ["序号", "原始文件名", "显示名称", "状态", "备注", "结果路径"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        for index, item in enumerate(processed, start=2):
            sheet.cell(row=index, column=1, value=index - 1)
            sheet.cell(row=index, column=2, value=item["原始文件名"])
            sheet.cell(row=index, column=3, value=item["显示名称"])
            sheet.cell(row=index, column=4, value=item["状态"])
            sheet.cell(row=index, column=5, value=item["备注"])
            sheet.cell(row=index, column=6, value=item["结果路径"])
        上报进度(progress, 90, "写入报告", "正在输出处理报告。")
        workbook.save(output_excel)
        print(f"处理报告已输出: {output_excel}")
        上报进度(progress, 100, "已完成", "Excel 映射重命名完成。")
        return {
            "processed_count": len(processed),
            "output_excel": str(output_excel),
            "renamed_count": sum(1 for item in processed if item["状态"] == "已重命名"),
            "pending_count": sum(1 for item in processed if item["状态"] == "待重命名"),
        }

    result, logs = 捕获日志(runner)
    return 返回成功("Excel 映射重命名已完成", logs, result)


stats_analysis = UP主统计分析
keyword_sort = 关键词整理
excel_rename = Excel映射重命名
